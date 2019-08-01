'''
Created on Jul 30, 2019

@author: naveen_pj
'''
from tkinter import *
import tkinter as tk
import openpyxl
from tkinter import messagebox

#open Excelsheet
work_book = openpyxl.load_workbook("/home/naveen_pj/Videos/register_form.xlsx")

#activate the excel sheet
sheet=work_book.active

#DEFINE EXCEL SIZE
def define_excel():
    sheet.column_dimensions['A'].width=20
    sheet.column_dimensions['B'].width=20
    sheet.column_dimensions['C'].width=20
    sheet.column_dimensions['D'].width=20
    sheet.column_dimensions['E'].width=20
    sheet.column_dimensions['F'].width=20
    sheet.column_dimensions['G'].width=20
    sheet.column_dimensions['H'].width=20
    sheet.column_dimensions['I'].width=20
    sheet.column_dimensions['J'].width=20
    sheet.column_dimensions['K'].width=20
    sheet.column_dimensions['L'].width=20
    
    #DEFINE EXCEL HEADING VALUE 
    sheet.cell(row=1 , column=1).value= "First Name"
    sheet.cell(row=1 , column=2).value= "Last Name"
    sheet.cell(row=1 , column=3).value= "destination"
    sheet.cell(row=1 , column=4).value= "year of experience"
    sheet.cell(row=1 , column=5).value= "Address"
    sheet.cell(row=1 , column=6).value= "city"
    sheet.cell(row=1 , column=7).value= "district"
    sheet.cell(row=1 , column=8).value= "State"
    sheet.cell(row=1 , column=9).value= "Pincode"
    sheet.cell(row=1 , column=10).value= "Mobile Number"
    sheet.cell(row=1 , column=11).value= "Email address" 

#Function to focus on get events
def focus1(event):
    first_name.focus_set()
    
def focus2(event):
    last_name.focus_set()
    
def focus3(event):
    destination.focus_set()

def focus4(event):
    expr.focus_set()
       
def focus5(event):
    email.focus_set()

def focus6(event):
    address.focus_set()
    
def focus7(event):
    state.focus_set()
    
def focus8(event):
    city.focus_set()
    
def focus9(event):
    pincode.focus_set()
    
def focus10(event):
    district.focus_set()

def focus11(event):
    mob_num.focus_set()
    
#function for clear the data in the form
def clear():
    first_name.delete(0, END)
    last_name.delete(0,END)
    destination.delete(0,END)
    expr.delete(0,END)
    email.delete(0, END)
    state.delete(0, END)
    city.delete(0, END)
    district.delete(0, END)
    address.delete(0, END)
    pincode.delete(0, END)
    mob_num.delete(0, END)

#Excel function inserting informations
def insert_details():
    
    if (first_name.get()=="" and last_name.get()=="" and destination.get()=="" and expr.get()=="" and 
        email.get()=="" and mob_num.get()=="" and mob_num.get()=="" and city.get()=="" ):
            tk.messagebox.showinfo("Error", "Please enter your details")
    else:
        current_row=sheet.max_row
        current_column=sheet.max_column
        
        #store the details from tkinder
        #excel spreadsheet at particular location
        sheet.cell(row=current_row + 1, column=1).value = first_name.get()
        sheet.cell(row=current_row + 1, column=2).value = last_name.get()
        sheet.cell(row=current_row + 1, column=3).value = destination.get()
        sheet.cell(row=current_row + 1, column=4).value = expr.get()
        sheet.cell(row=current_row + 1, column=5).value = address.get()
        sheet.cell(row=current_row + 1, column=6).value = city.get()
        sheet.cell(row=current_row + 1, column=7).value = district.get()
        sheet.cell(row=current_row + 1, column=8).value = state.get()
        sheet.cell(row=current_row + 1, column=9).value = pincode.get()
        sheet.cell(row=current_row + 1, column=10).value = mob_num.get()
        sheet.cell(row=current_row + 1, column=11).value = email.get()
        
        work_book.save("/home/naveen_pj/Videos/register_form.xlsx")
        
        clear()
        tk.messagebox.showinfo("Registration", "Registered sucessfully")


#drive code
if __name__ == "__main__": 
    
  # create a GUI window 
    root = Tk() 
  
    # set the background colour of GUI window 
    root.configure(background='light green') 
  
    # set the title of GUI window 
    root.title("Registration form") 
  
    # set the configuration of GUI window 
    root.geometry("500x320") 
    
    define_excel()
    # create a Form label 
    heading = Label(root, text="Registration Form", bg="light green") 
  
    # create a Name label 
    first_name = Label(root, text="First Name", bg="light green") 
    last_name = Label(root, text="Last Name", bg="light green")
  
    # create a destination label 
    destination = Label(root, text="Destination", bg="light green") 
  
    # create a experience label 
    expr = Label(root, text="year of experience", bg="light green") 
 
    # create a address id label 
    address = Label(root, text="Address", bg="light green") 
  
    # create a city label 
    city = Label(root, text="City", bg="light green") 
    
    # create a district label 
    district = Label(root, text="District", bg="light green") 
    
    # create a state label 
    state = Label(root, text="State", bg="light green") 
    
    # create a pincode label 
    pincode = Label(root, text="Pincode", bg="light green") 
    
    # create a mobile number label 
    mob_num = Label(root, text="Mobile Number", bg="light green") 
  
    # create a email label 
    email = Label(root, text="Email Id", bg="light green") 
    
    # grid method is used for placing 
    # the widgets at respective positions 
    # in table like structure . 
    heading.grid(row=0, column=1) 
    first_name.grid(row=1, column=0)
    last_name.grid(row=2, column=0)
    destination.grid(row=3, column=0) 
    expr.grid(row=4, column=0)
    address.grid(row=5, column=0) 
    city.grid(row=6, column=0) 
    district.grid(row=7, column=0)
    state.grid(row=8, column=0)
    pincode.grid(row=9, column=0)
    mob_num.grid(row=10, column=0)
    email.grid(row=11, column=0)
   
    
    # create a text entry box 
    # for typing the information 
    first_name = Entry(root) 
    last_name = Entry(root) 
    destination = Entry(root) 
    expr = Entry(root) 
    address = Entry(root) 
    city = Entry(root) 
    district = Entry(root)
    state = Entry(root)
    pincode = Entry(root)
    mob_num = Entry(root)
    email = Entry(root)
    

    # the binding the function with the events 
    # Fuction for focusing the input field 
    first_name.bind("<Return>", focus1) 
    last_name.bind("<Return>", focus2)  
    destination.bind("<Return>", focus3) 
    expr.bind("<Return>", focus4) 
    email.bind("<Return>", focus6) 
    address.bind("<Return>", focus5)
    state.bind("<Return>", focus8) 
    city.bind("<Return>", focus9) 
    pincode.bind("<Return>", focus10) 
    mob_num.bind("<Return>", focus11) 

  
    # grid method is used for placing the fields
    # the widgets at respective positions 
    # in table like structure . 
    first_name.grid(row=1, column=1, ipadx="100") 
    last_name.grid(row=2, column=1, ipadx="100") 
    destination.grid(row=3, column=1, ipadx="100") 
    expr.grid(row=4, column=1, ipadx="100")
    address.grid(row=5, column=1, ipadx="100") 
    city.grid(row=6, column=1, ipadx="100")
    district.grid(row=7, column=1, ipadx="100")
    state.grid(row=8, column=1, ipadx="100")
    pincode.grid(row=9, column=1, ipadx="100")
    mob_num.grid(row=10, column=1, ipadx="100")
    email.grid(row=11, column=1, ipadx="100")
  
    # calling excel function 
    define_excel() 
  
    # create a Submit Button and place into the root window 
    submit = Button(root, text="Submit", fg="Black", 
                            bg="Red", command=insert_details) 
    submit.grid(row=12, column=1) 
  
    # runing the GUI 
    root.mainloop() 
    
